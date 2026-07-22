"""Excel application tracker (applications.xlsx) via openpyxl."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.settings_loader import load_settings, resolve_path

TRACKER_COLUMNS = [
    "Date",
    "Company",
    "Role",
    "Job Link",
    "Status",
    "Resume File",
    "Cover Letter File",
    "Email File",
    "Notes",
]

_DEFAULT_STATUS = "Applied"


class ApplicationTracker:
    """Thin wrapper around applications.xlsx."""

    def __init__(self, path: Path | str) -> None:
        self.path = Path(path)

    def ensure_tracker(self) -> Path:
        return ensure_tracker(self.path)

    def add_application(
        self,
        company: str,
        role: str,
        job_link: str = "",
        status: str = _DEFAULT_STATUS,
        resume_file: str = "",
        cover_letter_file: str = "",
        email_file: str = "",
        notes: str = "",
        applied_date: str | None = None,
    ) -> int:
        return add_application(
            self.path,
            company=company,
            role=role,
            job_link=job_link,
            status=status,
            resume_file=resume_file,
            cover_letter_file=cover_letter_file,
            email_file=email_file,
            notes=notes,
            applied_date=applied_date,
        )

    def update_status(self, row_or_company: int | str, status: str) -> bool:
        return update_status(self.path, row_or_company, status)

    def list_applications(self) -> list[dict[str, Any]]:
        return list_applications(self.path)


def ensure_tracker(path: Path | str) -> Path:
    """Create tracker workbook with headers if missing."""
    tracker_path = Path(path)
    tracker_path.parent.mkdir(parents=True, exist_ok=True)

    if tracker_path.exists():
        wb = load_workbook(tracker_path)
        ws = wb.active
        _ensure_headers(ws)
        wb.save(tracker_path)
        return tracker_path

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.append(TRACKER_COLUMNS)
    wb.save(tracker_path)
    return tracker_path


def _ensure_headers(ws: Worksheet) -> None:
    existing = [cell.value for cell in ws[1]]
    if not existing or all(v is None for v in existing):
        for idx, header in enumerate(TRACKER_COLUMNS, start=1):
            ws.cell(row=1, column=idx, value=header)
        return

    for col_idx, header in enumerate(TRACKER_COLUMNS, start=1):
        if col_idx > len(existing) or existing[col_idx - 1] != header:
            ws.cell(row=1, column=col_idx, value=header)


def _today_str() -> str:
    return date.today().isoformat()


def add_application(
    path: Path | str,
    company: str,
    role: str,
    job_link: str = "",
    status: str = _DEFAULT_STATUS,
    resume_file: str = "",
    cover_letter_file: str = "",
    email_file: str = "",
    notes: str = "",
    applied_date: str | None = None,
) -> int:
    """Append a new application row. Returns the 1-based Excel row number."""
    tracker_path = ensure_tracker(path)
    wb = load_workbook(tracker_path)
    ws = wb.active

    row = [
        applied_date or _today_str(),
        company,
        role,
        job_link,
        status or _DEFAULT_STATUS,
        resume_file,
        cover_letter_file,
        email_file,
        notes,
    ]
    ws.append(row)
    row_num = ws.max_row
    wb.save(tracker_path)
    return row_num


def update_status(path: Path | str, row_or_company: int | str, status: str) -> bool:
    """Update Status by 1-based data row number (row 2 = first application) or company name."""
    tracker_path = ensure_tracker(path)
    wb = load_workbook(tracker_path)
    ws = wb.active
    status_col = TRACKER_COLUMNS.index("Status") + 1

    if isinstance(row_or_company, int):
        target_row = row_or_company
        if target_row < 2 or target_row > ws.max_row:
            return False
        ws.cell(row=target_row, column=status_col, value=status)
        wb.save(tracker_path)
        return True

    company_col = TRACKER_COLUMNS.index("Company") + 1
    needle = str(row_or_company).strip().lower()
    updated = False
    for row_idx in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=company_col).value
        if cell_val and str(cell_val).strip().lower() == needle:
            ws.cell(row=row_idx, column=status_col, value=status)
            updated = True
    if updated:
        wb.save(tracker_path)
    return updated


def list_applications(path: Path | str) -> list[dict[str, Any]]:
    """Return all application rows as dicts (most recent last)."""
    tracker_path = ensure_tracker(path)
    wb = load_workbook(tracker_path, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter, None)
    if not headers:
        wb.close()
        return []

    header_list = [str(h) if h is not None else "" for h in headers]
    records: list[dict[str, Any]] = []
    for excel_row in rows_iter:
        if not excel_row or all(v is None or v == "" for v in excel_row):
            continue
        record: dict[str, Any] = {"_row": len(records) + 2}
        for idx, header in enumerate(header_list):
            if not header:
                continue
            value = excel_row[idx] if idx < len(excel_row) else None
            if isinstance(value, datetime):
                value = value.date().isoformat()
            record[header] = value
        records.append(record)

    wb.close()
    return records


def default_tracker_path(settings: dict[str, Any] | None = None) -> Path:
    cfg = settings or load_settings()
    return resolve_path("tracker_xlsx", cfg)
